import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import NamedStyle

# Função para ler o XML
def parse_xml(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    return root

# Função para extrair os dados de interesse do XML
def extract_data_from_xml(root):
    records = []

    for cp in root.findall(".//ns2:CprDhConsultar", namespaces={"ns2": "http://services.docHabil.cpr.siafi.tesouro.fazenda.gov.br/"}):
        ano_dh = cp.find(".//anoDH").text if cp.find(".//anoDH") is not None else ''
        cod_tipo_dh = cp.find(".//codTipoDH").text if cp.find(".//codTipoDH") is not None else ''
        num_dh = cp.find(".//numDH").text if cp.find(".//numDH") is not None else ''
        data_fato_gerador = cp.find(".//Data_fato_gerador").text if cp.find(".//Data_fato_gerador") is not None else ''
        ob = cp.find(".//OB").text if cp.find(".//OB") is not None else ''
        valor_ob = cp.find(".//VALOR_OB").text if cp.find(".//VALOR_OB") is not None else ''
        cod_sit = cp.find(".//codSit").text if cp.find(".//codSit") is not None else ''
        txt_inscr_a = cp.find(".//txtInscrA").text if cp.find(".//txtInscrA") is not None else ''
        txt_inscr_b = cp.find(".//txtInscrB").text if cp.find(".//txtInscrB") is not None else ''
        cod_recolhedor = cp.find(".//codRecolhedor").text if cp.find(".//codRecolhedor") is not None else ''

        # Valores de retenção
        vlr = cp.find(".//itemRecolhimento/vlr").text if cp.find(".//itemRecolhimento/vlr") is not None else ''
        vlr_base_calculo = cp.find(".//vlrBaseCalculo").text if cp.find(".//vlrBaseCalculo") is not None else ''
        vlr_multa = cp.find(".//vlrMulta").text if cp.find(".//vlrMulta") is not None else ''
        vlr_juros = cp.find(".//vlrJuros").text if cp.find(".//vlrJuros") is not None else ''
        vlr_outras_ent = cp.find(".//vlrOutrasEnt").text if cp.find(".//vlrOutrasEnt") is not None else ''
        vlr_atm_multa_juros = cp.find(".//vlrAtmMultaJuros").text if cp.find(".//vlrAtmMultaJuros") is not None else ''

        predoc = cp.find(".//predoc")
        txt_obser = predoc.find(".//txtObser").text if predoc is not None and predoc.find(".//txtObser") is not None else ''

        # Verificar se pelo menos um valor de retenção está presente
        if any([vlr, vlr_multa, vlr_juros, vlr_outras_ent, vlr_atm_multa_juros]):
            records.append([ano_dh, cod_tipo_dh, num_dh, data_fato_gerador, ob, valor_ob, cod_sit,
                            txt_inscr_a, txt_inscr_b, cod_recolhedor, vlr, vlr_base_calculo,
                            vlr_multa, vlr_juros, vlr_outras_ent, vlr_atm_multa_juros, txt_obser])

    return records

# Função para criar e salvar a planilha Excel
def create_excel(data, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracao SIAFI"

    headers = ["Ano DH", "Código Tipo DH", "Num DH", "Data Fato Gerador", "OB", "Valor OB", "Código Sit",
               "Inscrição A", "Inscrição B", "Código Recolhedor", "Valor IR", "Valor Base Cálculo",
               "Valor Multa", "Valor Juros", "Valor Outras Ent", "Valor Atm Multa Juros", "Observação"]
    ws.append(headers)
    
    # Filtrar dados onde a coluna "Valor OB" (F) tem valor
    filtered_data = [record for record in data if record[5]]  # "Valor OB" presente

    # Criar um dicionário para agrupar primeiro por Código Recolhedor (Coluna J - índice 9) e depois por txtInscrB (Coluna I - índice 8)
    grupos = {}
    for record in filtered_data:
        cod_recolhedor = record[9]
        txt_inscr_b = record[8]
        
        if cod_recolhedor not in grupos:
            grupos[cod_recolhedor] = {}
        
        if txt_inscr_b not in grupos[cod_recolhedor]:
            grupos[cod_recolhedor][txt_inscr_b] = []
        
        grupos[cod_recolhedor][txt_inscr_b].append(record)

    # Criar estilo contábil
    if "Contábil" not in wb.named_styles:
        contabilidade = NamedStyle(name="Contábil", number_format="#,##0.00")
        wb.add_named_style(contabilidade)
    
    # Preencher planilha agrupada
    for cod_recolhedor, inscritos_b in grupos.items():
        for txt_inscr_b, registros in inscritos_b.items():
            total_f = total_k = total_l = total_m = total_n = total_o = total_p = 0
            
            # Adicionar registros de cada grupo e somar os valores
            for record in registros:
                # Converter valores numéricos corretamente e somar
                valor_f = float(record[5].replace(",", ".") if record[5] else 0)
                valor_k = float(record[10].replace(",", ".") if record[10] else 0)
                valor_l = float(record[11].replace(",", ".") if record[11] else 0)
                valor_m = float(record[12].replace(",", ".") if record[12] else 0)
                valor_n = float(record[13].replace(",", ".") if record[13] else 0)
                valor_o = float(record[14].replace(",", ".") if record[14] else 0)
                valor_p = float(record[15].replace(",", ".") if record[15] else 0)

                total_f += valor_f
                total_k += valor_k
                total_l += valor_l
                total_m += valor_m
                total_n += valor_n
                total_o += valor_o
                total_p += valor_p

                ws.append(record)
            
            # Adicionar linha de soma abaixo do grupo txtInscrB
            def formatar_valor(valor):
                valor_float = float(valor)
                return f"{valor_float:.2f}".replace(".", ",")       
            
            total_row = ["TOTAL POR ND POR CNPJ:"]+[""] * 4 + [formatar_valor(total_f)] + [""] * 2 + [txt_inscr_b,cod_recolhedor] + [
                formatar_valor(total_k), formatar_valor(total_l), formatar_valor(total_m),
                formatar_valor(total_n), formatar_valor(total_o), formatar_valor(total_p),
                "RECOLHIMENTO DE TRIBUTOS FEDERAIS, SOBRE 2 OU MAIS NF-Es, EM ATENDIMENTO A IN/RFB 1234/2012."
            ]
            
            # Adicionar a linha formatada na planilha
            ws.append(total_row)
        
    # Formatar as células
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=16):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.replace(".", "").isdigit():
                cell.value = cell.value.replace(".", ",")
            cell.style = "Contábil"

    # Salvar a planilha
    wb.save(output_file)

# Função principal
def main():
    xml_file = "ArquivosUnidos/Consolidado_Siafi_Tesouro.xml"
    output_file = "RelatorioGerado/Relatorio.xlsx"

    root = parse_xml(xml_file)
    extracted_data = extract_data_from_xml(root)
    create_excel(extracted_data, output_file)

    print(f"Arquivo {output_file} criado com sucesso!")

if __name__ == "__main__":
    main()
